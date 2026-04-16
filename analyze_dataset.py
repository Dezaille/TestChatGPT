#!/usr/bin/env python3
"""Analyse un fichier CSV/XLSX et produit une description sans exposer les données."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from datetime import datetime
from pathlib import Path
from typing import Any

SUPPORTED_EXTENSIONS = {".csv", ".xlsx"}
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
URL_RE = re.compile(r"^https?://")


def is_null(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def try_parse_number(value: str) -> float | None:
    candidate = value.strip().replace(" ", "")
    candidate = candidate.replace(",", ".")
    try:
        return float(candidate)
    except ValueError:
        return None


def try_parse_date(value: str) -> bool:
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"]
    for fmt in formats:
        try:
            datetime.strptime(value.strip(), fmt)
            return True
        except ValueError:
            continue
    return False


def read_csv_rows(path: Path, encoding: str, separator: str) -> tuple[list[str], list[list[Any]]]:
    with path.open("r", encoding=encoding, newline="") as fh:
        reader = csv.reader(fh, delimiter=separator)
        rows = list(reader)

    if not rows:
        return [], []
    return rows[0], rows[1:]


def read_xlsx_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Le support XLSX nécessite openpyxl. Installez-le avec: pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        return [], []

    header = ["" if cell is None else str(cell) for cell in rows[0]]
    data = [list(r) for r in rows[1:]]
    return header, data


def classify_column(values: list[Any]) -> tuple[str, str]:
    non_null = [v for v in values if not is_null(v)]
    if not non_null:
        return "empty", "texte"

    numeric_values: list[float] = []
    bool_like = 0
    date_like = 0
    email_like = 0
    url_like = 0

    for value in non_null:
        if isinstance(value, bool):
            bool_like += 1
            continue
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            numeric_values.append(float(value))
            continue

        string_value = str(value).strip()
        lowered = string_value.lower()
        if lowered in {"true", "false", "oui", "non", "0", "1"}:
            bool_like += 1

        parsed_number = try_parse_number(string_value)
        if parsed_number is not None and not math.isnan(parsed_number):
            numeric_values.append(parsed_number)

        if try_parse_date(string_value):
            date_like += 1
        if EMAIL_RE.match(string_value):
            email_like += 1
        if URL_RE.match(string_value):
            url_like += 1

    total = len(non_null)
    numeric_ratio = len(numeric_values) / total
    bool_ratio = bool_like / total
    date_ratio = date_like / total
    email_ratio = email_like / total
    url_ratio = url_like / total

    if bool_ratio > 0.9:
        return "bool", "booléen"
    if numeric_ratio > 0.9:
        is_integer = all(float(num).is_integer() for num in numeric_values)
        return ("int" if is_integer else "float", "entier" if is_integer else "nombre décimal")
    if date_ratio > 0.8:
        return "date", "date potentielle"
    if email_ratio > 0.8:
        return "string", "email"
    if url_ratio > 0.8:
        return "string", "URL"

    return "string", "texte"


def describe_columns(header: list[str], rows: list[list[Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    row_count = len(rows)

    for idx, name in enumerate(header):
        col_name = name or f"colonne_{idx + 1}"
        values = [row[idx] if idx < len(row) else None for row in rows]
        null_count = sum(1 for v in values if is_null(v))
        dtype, semantic_type = classify_column(values)

        column_meta: dict[str, Any] = {
            "nom": col_name,
            "dtype_inferé": dtype,
            "type_semantique": semantic_type,
            "valeurs_vides_possibles": null_count > 0,
            "description_contenu_colonne": "",
        }

        result.append(column_meta)

    return result


def describe_dataset(path: Path, encoding: str = "utf-8", separator: str = ",") -> dict[str, Any]:
    ext = path.suffix.lower()
    if ext == ".csv":
        header, rows = read_csv_rows(path, encoding, separator)
    elif ext == ".xlsx":
        header, rows = read_xlsx_rows(path)
    else:
        raise ValueError("Format non supporté. Utilisez .csv ou .xlsx")

    if not header:
        return {
            "fichier": path.name,
            "format": ext.lstrip("."),
            "nombre_lignes": 0,
            "nombre_colonnes": 0,
            "colonnes": [],
            "note_confidentialite": "Fichier vide ou sans en-tête détectable.",
        }

    columns = describe_columns(header, rows)

    return {
        "fichier": path.name,
        "format": ext.lstrip("."),
        "nombre_lignes": len(rows),
        "nombre_colonnes": len(header),
        "colonnes": columns,
        "note_confidentialite": (
            "Ce rapport ne contient pas de valeurs brutes; uniquement une description "
            "de la structure et des types de colonnes."
        ),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Analyse un fichier CSV/XLSX et génère une description exploitable par une IA."
    )
    parser.add_argument("input_file", type=Path, help="Chemin vers le fichier .csv ou .xlsx")
    parser.add_argument("-o", "--output", type=Path, default=Path("dataset_description.json"))
    parser.add_argument("--encoding", default="utf-8", help="Encodage CSV (défaut: utf-8)")
    parser.add_argument("--separator", default=",", help="Séparateur CSV (défaut: ,)")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.input_file.exists():
        raise SystemExit(f"Le fichier '{args.input_file}' est introuvable.")
    if args.input_file.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise SystemExit("Format non supporté. Utilisez .csv ou .xlsx")

    description = describe_dataset(args.input_file, encoding=args.encoding, separator=args.separator)
    args.output.write_text(json.dumps(description, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Description générée: {args.output}")


if __name__ == "__main__":
    main()
